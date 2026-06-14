"""
Document service: file upload, text extraction, chunking, embedding, and Qdrant storage.
Uses BGE-M3 for embeddings.
"""
import hashlib
import os
import uuid
from pathlib import Path
from typing import List, Optional, Tuple, Dict, Any

import aiofiles
import structlog
from fastapi import HTTPException, UploadFile, status
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import settings
from app.models.document import Document, DocumentStatus, DocumentType
from app.models.user import User
from app.repositories.document import DocumentRepository
from app.utils.file_handler import FileHandler
from app.utils.text_processor import TextProcessor

logger = structlog.get_logger(__name__)


class DocumentService:
    """
    Service for document upload, processing, and embedding.
    Handles the full pipeline: upload → extract → chunk → embed → store in Qdrant.
    """

    def __init__(self, db: AsyncSession) -> None:
        self.db = db
        self.doc_repo = DocumentRepository(db)
        self.file_handler = FileHandler()
        self.text_processor = TextProcessor()
        self._embedder = None
        self._qdrant_client = None

    def _get_embedder(self):
        """Lazy-load BGE-M3 embedder."""
        if self._embedder is None:
            try:
                from FlagEmbedding import BGEM3FlagModel
                self._embedder = BGEM3FlagModel(
                    settings.bge_model_name,
                    use_fp16=True,
                    device=settings.embedding_device,
                )
                logger.info("BGE-M3 embedder loaded", model=settings.bge_model_name)
            except ImportError:
                logger.warning("FlagEmbedding not available, falling back to OpenAI embeddings")
                self._embedder = "openai"
        return self._embedder

    def _get_qdrant_client(self):
        """Lazy-load Qdrant client."""
        if self._qdrant_client is None:
            from qdrant_client import QdrantClient
            self._qdrant_client = QdrantClient(
                host=settings.qdrant_host,
                port=settings.qdrant_port,
                api_key=settings.qdrant_api_key or None,
            )
        return self._qdrant_client

    async def upload_document(
        self,
        file: UploadFile,
        user: User,
    ) -> Document:
        """
        Validate and save an uploaded document, then trigger async processing.

        Args:
            file: The uploaded file
            user: The authenticated user

        Returns:
            Created Document record

        Raises:
            HTTPException 400: If file type or size is invalid
        """
        # Validate file
        await self.file_handler.validate_file(file)

        # Save file to disk
        file_path, stored_filename, file_size = await self.file_handler.save_file(
            file, user_id=str(user.id)
        )

        # Determine file type
        ext = Path(file.filename).suffix.lstrip(".").lower()
        try:
            file_type = DocumentType(ext)
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Unsupported file type: {ext}",
            )

        # Compute content hash
        content_hash = await self._compute_file_hash(file_path)

        # Check for duplicate
        existing = await self.doc_repo.get_by_content_hash(content_hash, user.id)
        if existing:
            # Remove the duplicate upload
            Path(file_path).unlink(missing_ok=True)
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="This document has already been uploaded.",
            )

        # Create database record
        document = await self.doc_repo.create({
            "user_id": user.id,
            "filename": stored_filename,
            "original_filename": file.filename,
            "file_path": file_path,
            "file_type": file_type,
            "file_size": file_size,
            "status": DocumentStatus.PENDING,
            "content_hash": content_hash,
        })

        logger.info(
            "Document uploaded",
            document_id=str(document.id),
            filename=file.filename,
            user_id=str(user.id),
        )

        # Trigger async processing via Celery (if available)
        try:
            from app.tasks.document_tasks import process_document_task
            process_document_task.delay(str(document.id))
        except ImportError:
            # If Celery is not set up, process synchronously
            logger.warning("Celery not available, processing document synchronously")
            await self.process_document(str(document.id))

        return document

    async def process_document(self, document_id: str) -> None:
        """
        Full document processing pipeline:
        1. Extract text from document
        2. Chunk text
        3. Generate BGE-M3 embeddings
        4. Store in Qdrant

        Args:
            document_id: UUID string of the document to process
        """
        doc_uuid = uuid.UUID(document_id)
        document = await self.doc_repo.get_by_id(doc_uuid)

        if not document:
            logger.error("Document not found for processing", document_id=document_id)
            return

        try:
            # Update status to processing
            await self.doc_repo.update_status(doc_uuid, DocumentStatus.PROCESSING)

            # Step 1: Extract text
            logger.info("Extracting text", document_id=document_id)
            text, page_count = await self._extract_text(document)

            if not text or not text.strip():
                raise ValueError("No text could be extracted from the document")

            # Store extracted text
            await self.doc_repo.update(doc_uuid, {
                "extracted_text": text[:50000],  # Store first 50k chars
                "page_count": page_count,
            })

            # Step 2: Chunk text
            logger.info("Chunking text", document_id=document_id)
            chunks = self.text_processor.chunk_text(
                text=text,
                chunk_size=settings.chunk_size,
                chunk_overlap=settings.chunk_overlap,
                document_id=document_id,
                filename=document.original_filename,
            )

            # Step 3: Generate embeddings and store in Qdrant
            if chunks and settings.enable_rag:
                logger.info(
                    "Embedding and storing chunks",
                    document_id=document_id,
                    chunk_count=len(chunks),
                )
                await self._embed_and_store(document, chunks)

            # Step 4: Update document record
            await self.doc_repo.update_processing_result(
                document_id=doc_uuid,
                chunk_count=len(chunks),
                page_count=page_count,
                qdrant_collection=settings.qdrant_collection_name,
            )

            logger.info(
                "Document processed successfully",
                document_id=document_id,
                chunks=len(chunks),
                pages=page_count,
            )

        except Exception as e:
            logger.error(
                "Document processing failed",
                document_id=document_id,
                error=str(e),
                exc_info=True,
            )
            await self.doc_repo.update_status(
                doc_uuid,
                DocumentStatus.FAILED,
                error_message=str(e)[:500],
            )

    async def _extract_text(self, document: Document) -> Tuple[str, int]:
        """Extract text from a document based on its file type."""
        file_path = document.file_path
        file_type = document.file_type.value

        try:
            if file_type == "pdf":
                return await self._extract_from_pdf(file_path)
            elif file_type in ("docx", "doc"):
                return await self._extract_from_docx(file_path)
            elif file_type == "txt":
                return await self._extract_from_txt(file_path)
            elif file_type in ("xlsx", "xls"):
                return await self._extract_from_excel(file_path)
            else:
                raise ValueError(f"Unsupported file type for extraction: {file_type}")
        except Exception as e:
            logger.error("Text extraction failed", file_type=file_type, error=str(e))
            raise

    async def _extract_from_pdf(self, file_path: str) -> Tuple[str, int]:
        """Extract text from PDF using pypdf."""
        from pypdf import PdfReader
        reader = PdfReader(file_path)
        pages = []
        for page in reader.pages:
            text = page.extract_text()
            if text:
                pages.append(text)
        return "\n\n".join(pages), len(reader.pages)

    async def _extract_from_docx(self, file_path: str) -> Tuple[str, int]:
        """Extract text from DOCX using python-docx."""
        from docx import Document as DocxDocument
        doc = DocxDocument(file_path)
        paragraphs = [para.text for para in doc.paragraphs if para.text.strip()]
        return "\n\n".join(paragraphs), max(1, len(paragraphs) // 40)

    async def _extract_from_txt(self, file_path: str) -> Tuple[str, int]:
        """Extract text from plain text file."""
        async with aiofiles.open(file_path, mode="r", encoding="utf-8", errors="ignore") as f:
            content = await f.read()
        lines = content.count('\n')
        pages = max(1, lines // 50)
        return content, pages

    async def _extract_from_excel(self, file_path: str) -> Tuple[str, int]:
        """Extract text from Excel files."""
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True)
        sheets_text = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            rows_text = []
            for row in sheet.iter_rows(values_only=True):
                row_data = [str(cell) for cell in row if cell is not None]
                if row_data:
                    rows_text.append(" | ".join(row_data))
            if rows_text:
                sheets_text.append(f"Sheet: {sheet_name}\n" + "\n".join(rows_text))
        wb.close()
        return "\n\n".join(sheets_text), len(wb.sheetnames)

    async def _embed_and_store(
        self, document: Document, chunks: List[Dict[str, Any]]
    ) -> None:
        """Generate embeddings for chunks and store in Qdrant."""
        try:
            qdrant = self._get_qdrant_client()
            from qdrant_client.models import (
                Distance, VectorParams, PointStruct, SparseVector,
                SparseVectorParams, SparseIndexParams
            )

            # Ensure collection exists
            await self._ensure_qdrant_collection(qdrant)

            # Generate embeddings in batches
            texts = [chunk["text"] for chunk in chunks]
            embedder = self._get_embedder()

            points = []
            batch_size = settings.embedding_batch_size

            for batch_start in range(0, len(texts), batch_size):
                batch_texts = texts[batch_start:batch_start + batch_size]
                batch_chunks = chunks[batch_start:batch_start + batch_size]

                if embedder == "openai":
                    # Fallback to OpenAI embeddings
                    embeddings = await self._embed_with_openai(batch_texts)
                    for idx, (chunk, embedding) in enumerate(zip(batch_chunks, embeddings)):
                        point = PointStruct(
                            id=str(uuid.uuid4()),
                            vector={"dense": embedding},
                            payload={
                                "document_id": str(document.id),
                                "user_id": str(document.user_id),
                                "text": chunk["text"],
                                "chunk_index": chunk.get("chunk_index", batch_start + idx),
                                "page_number": chunk.get("page_number"),
                                "filename": document.original_filename,
                            }
                        )
                        points.append(point)
                else:
                    # Use BGE-M3 for dense + sparse (hybrid) embeddings
                    output = embedder.encode(
                        batch_texts,
                        return_dense=True,
                        return_sparse=True,
                        return_colbert_vecs=False,
                        batch_size=batch_size,
                    )

                    for idx, (chunk, dense_vec) in enumerate(
                        zip(batch_chunks, output["dense_vecs"])
                    ):
                        sparse_data = output["lexical_weights"][idx]
                        sparse_indices = [int(k) for k in sparse_data.keys()]
                        sparse_values = [float(v) for v in sparse_data.values()]

                        point = PointStruct(
                            id=str(uuid.uuid4()),
                            vector={
                                "dense": dense_vec.tolist(),
                                "sparse": SparseVector(
                                    indices=sparse_indices,
                                    values=sparse_values,
                                ),
                            },
                            payload={
                                "document_id": str(document.id),
                                "user_id": str(document.user_id),
                                "text": chunk["text"],
                                "chunk_index": chunk.get("chunk_index", batch_start + idx),
                                "page_number": chunk.get("page_number"),
                                "filename": document.original_filename,
                            }
                        )
                        points.append(point)

            # Upload to Qdrant in batches
            upload_batch_size = 100
            for i in range(0, len(points), upload_batch_size):
                batch = points[i:i + upload_batch_size]
                qdrant.upsert(
                    collection_name=settings.qdrant_collection_name,
                    points=batch,
                )

            logger.info(
                "Stored embeddings in Qdrant",
                document_id=str(document.id),
                points_count=len(points),
            )

        except Exception as e:
            logger.error("Failed to store embeddings", error=str(e), exc_info=True)
            raise

    async def _ensure_qdrant_collection(self, qdrant) -> None:
        """Create Qdrant collection if it doesn't exist."""
        from qdrant_client.models import (
            Distance, VectorParams, SparseVectorParams,
            SparseIndexParams, OptimizersConfigDiff
        )

        try:
            qdrant.get_collection(settings.qdrant_collection_name)
        except Exception:
            # Collection doesn't exist, create it
            embedder = self._get_embedder()
            vector_size = 1024  # BGE-M3 dense vector size

            if embedder == "openai":
                vector_size = 1536  # OpenAI text-embedding-3-small

                qdrant.create_collection(
                    collection_name=settings.qdrant_collection_name,
                    vectors_config={"dense": VectorParams(
                        size=vector_size,
                        distance=Distance.COSINE,
                    )},
                )
            else:
                qdrant.create_collection(
                    collection_name=settings.qdrant_collection_name,
                    vectors_config={"dense": VectorParams(
                        size=vector_size,
                        distance=Distance.COSINE,
                    )},
                    sparse_vectors_config={
                        "sparse": SparseVectorParams(
                            index=SparseIndexParams(on_disk=False)
                        )
                    },
                )

            logger.info(
                "Created Qdrant collection",
                collection=settings.qdrant_collection_name,
            )

    async def _embed_with_openai(self, texts: List[str]) -> List[List[float]]:
        """Fallback: generate embeddings using OpenAI."""
        from openai import AsyncOpenAI
        client = AsyncOpenAI(api_key=settings.openai_api_key)
        response = await client.embeddings.create(
            model=settings.openai_embedding_model,
            input=texts,
        )
        return [item.embedding for item in response.data]

    async def _compute_file_hash(self, file_path: str) -> str:
        """Compute SHA-256 hash of a file for deduplication."""
        sha256 = hashlib.sha256()
        async with aiofiles.open(file_path, "rb") as f:
            while chunk := await f.read(65536):
                sha256.update(chunk)
        return sha256.hexdigest()

    async def delete_document(
        self, document_id: uuid.UUID, user_id: uuid.UUID
    ) -> bool:
        """
        Soft-delete a document and remove its Qdrant embeddings.

        Returns:
            True if deleted, False if not found
        """
        document = await self.doc_repo.get_by_id_and_user(document_id, user_id)
        if not document:
            return False

        # Remove from Qdrant
        if document.status == DocumentStatus.PROCESSED:
            try:
                qdrant = self._get_qdrant_client()
                from qdrant_client.models import Filter, FieldCondition, MatchValue
                qdrant.delete(
                    collection_name=settings.qdrant_collection_name,
                    points_selector=Filter(
                        must=[
                            FieldCondition(
                                key="document_id",
                                match=MatchValue(value=str(document_id)),
                            )
                        ]
                    ),
                )
            except Exception as e:
                logger.error("Failed to delete from Qdrant", error=str(e))

        # Soft delete in database
        await self.doc_repo.soft_delete(document_id, user_id)

        # Delete physical file
        try:
            if document.file_path and os.path.exists(document.file_path):
                os.unlink(document.file_path)
        except Exception as e:
            logger.error("Failed to delete file", error=str(e))

        logger.info("Document deleted", document_id=str(document_id), user_id=str(user_id))
        return True
